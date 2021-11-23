from io import StringIO, BytesIO
import pandas as pd
from liquer.state_types import StateType, register_state_type
from liquer.constants import mimetype_from_extension

from liquer.commands import command, first_command
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import TemporaryDirectory
from pathlib import Path
from liquer.context import get_context
from itertools import islice

class OpenpyxlWorkbookStateType(StateType):
    def identifier(self):
        return "openpyxl_workbook"

    def default_extension(self):
        return "xlsx"

    def is_type_of(self, data):
        return isinstance(data, Workbook)

    def as_bytes(self, data, extension=None):
        if extension is None:
            extension = self.default_extension()
        assert self.is_type_of(data)
        mimetype = mimetype_from_extension(extension)

        if extension in ("xlsx","xltx"):
            with TemporaryDirectory() as tmpdir:
                path = Path(tmpdir) / f"data.{extension}"
                data.save(str(path))
                b=path.read_bytes()
                return b, mimetype
        else:
            raise Exception(
                f"Serialization: file extension {extension} is not supported by openpyxl_workbook type."
            )

    def from_bytes(self, b: bytes, extension=None):
        if extension is None:
            extension = self.default_extension()
        f = BytesIO()
        f.write(b)
        f.seek(0)
        if extension in ("xlsx","xltx"):
            return load_workbook(f)
        raise Exception(
            f"Deserialization: file extension {extension} is not supported by openpyxl_workbook type."
        )

    def copy(self, data):
        return self.from_bytes(self.as_bytes(data)[0])

    def data_characteristics(self, data):
        return dict(description=f"Excel workbook with {len(data.sheetnames)} sheets.")
        

OPENPYXL_WORKBOOK_STATE_TYPE = OpenpyxlWorkbookStateType()
register_state_type(Workbook, OPENPYXL_WORKBOOK_STATE_TYPE)

@command
def workbook(data, index=True, header=True, context=None):
    """Convert bytes or a dataframe to a workbook"""
    context=get_context(context)
    if type(data)==bytes:
        context.info("Workbook from bytes")
        return OPENPYXL_WORKBOOK_STATE_TYPE.from_bytes(data)
    elif isinstance(data,pd.DataFrame):
        context.info("Workbook from pandas DataFrame")
        wb=Workbook()
        ws=wb.active
        for r in dataframe_to_rows(df, index=index, header=header):
            ws.append(r)
        return wb
    elif isinstance(data,Workbook):
        return data
    raise Exception(f"Unsupported workbook type: {type(data)}")

@command
def workbook_sheet_df(wb, sheet=None, context=None):
    """Extract a workbook sheet as a data-frame"""
    context = get_context(context)
    if type(wb)==bytes:
        wb = workbook(wb, context=context)
    if sheet in ("",None):
        context.info("Using active sheet")
        ws=wb.active
    else:
        ws=wb[sheet]
    try:
        i=int(sheet)
        sheet = wb.sheetnames[i]
        context.info(f"Using sheet {i} with name '{sheet}'")
    except:
        pass

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)    
    return df